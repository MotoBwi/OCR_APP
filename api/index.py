import os
import tempfile
import json
import logging
import time
from flask import Flask, request, jsonify
from werkzeug.utils import secure_filename
from flask_httpauth import HTTPTokenAuth
import pdfplumber
from pdf2image import convert_from_path
from PIL import Image
import cv2
import numpy as np
import io
import pandas as pd
try:
    from docx import Document
except ImportError:
    Document = None  # Handle case where python-docx is not installed
import openpyxl
import easyocr

app = Flask(__name__)
auth = HTTPTokenAuth(scheme='Bearer')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Configuration
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'txt', 'csv', 'xlsx', 'xls', 'jpg', 'jpeg', 'png'}
UPLOAD_FOLDER = tempfile.mkdtemp()
OUTPUT_FOLDER = os.path.join(os.getcwd(), 'extracted_data')
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB limit

# Initialize EasyOCR readers with GPU support
reader_en_hi = easyocr.Reader(['en', 'hi'], gpu=True)
reader_en_bn = easyocr.Reader(['en', 'bn'], gpu=True)
reader_en_ur = easyocr.Reader(['en', 'ur'], gpu=True)

# Define your API key
API_KEY = "adsgsdssHFGkh@sagsjdkwhdq"

@auth.verify_token
def verify_token(token):
    return token == API_KEY

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def preprocess_image(image):
    """Enhance image for better OCR results"""
    try:
        img = np.array(image)
        if len(img.shape) == 2:  # Grayscale
            img = cv2.cvtColor(img, cv2.COLOR_GRAY2RGB)
        elif img.shape[2] == 4:  # RGBA
            img = cv2.cvtColor(img, cv2.COLOR_RGBA2RGB)

        # Convert to grayscale for processing
        gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)

        # Apply adaptive thresholding
        processed = cv2.adaptiveThreshold(
            gray, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY, 11, 2
        )

        return Image.fromarray(processed)
    except Exception as e:
        logger.error(f"Image preprocessing failed: {str(e)}")
        return image

def extract_text_from_image(image):
    """Extract text from image using EasyOCR"""
    try:
        processed_img = preprocess_image(image)
        result_en_hi = reader_en_hi.readtext(np.array(processed_img))
        result_en_bn = reader_en_bn.readtext(np.array(processed_img))
        result_en_ur = reader_en_ur.readtext(np.array(processed_img))

        text_en_hi = " ".join([text[1] for text in result_en_hi])
        text_en_bn = " ".join([text[1] for text in result_en_bn])
        text_en_ur = " ".join([text[1] for text in result_en_ur])

        return text_en_hi + " " + text_en_bn + " " + text_en_ur
    except Exception as e:
        logger.error(f"OCR extraction failed: {str(e)}")
        return ""

def process_pdf_page(page, page_num, pdf_path):
    """Process a single PDF page with mixed content"""
    result = {
        "page": page_num + 1,
        "native_text": "",
        "image_text": "",
        "type": "mixed"
    }

    # First try to extract native text
    try:
        result["native_text"] = page.extract_text(x_tolerance=1, y_tolerance=1) or ""
    except Exception as e:
        logger.warning(f"Native text extraction failed: {str(e)}")

    # Check if page has images or if native text extraction was insufficient
    if page.images or len(result["native_text"].strip()) < 50:
        try:
            # Convert the entire page to image
            images = convert_from_path(
                pdf_path,
                first_page=page_num+1,
                last_page=page_num+1,
                dpi=300,
                size=(2480, 3508))  # A4 size at 300dpi

            if images:
                # Extract text from the full page image
                full_page_text = extract_text_from_image(images[0])

                # Only use OCR text if we got more content than native extraction
                if len(full_page_text) > len(result["native_text"]):
                    result["image_text"] = full_page_text
                    result["type"] = "ocr_text" if not result["native_text"] else "mixed"

                # Explicit cleanup
                del images
        except Exception as e:
            logger.error(f"Page image processing failed: {str(e)}")

    return result

def process_docx(file_path):
    """Extract text from DOCX file"""
    if Document is None:
        raise ImportError("python-docx package is not installed")

    try:
        doc = Document(file_path)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return {
            "content": [{
                "page": 1,
                "text": text,
                "type": "native_text"
            }]
        }
    except Exception as e:
        logger.error(f"DOCX processing failed: {str(e)}")
        raise

def process_txt(file_path):
    """Extract text from TXT file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            text = f.read()
        return {
            "content": [{
                "page": 1,
                "text": text,
                "type": "native_text"
            }]
        }
    except Exception as e:
        logger.error(f"TXT processing failed: {str(e)}")
        raise

def process_csv(file_path):
    """Extract data from CSV file"""
    try:
        df = pd.read_csv(file_path)
        text = df.to_string(index=False)
        return {
            "content": [{
                "page": 1,
                "text": text,
                "type": "table_data"
            }]
        }
    except Exception as e:
        logger.error(f"CSV processing failed: {str(e)}")
        raise

def process_excel(file_path):
    """Extract data from Excel file (XLSX or XLS)"""
    try:
        text = ""
        if file_path.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"\n\nSheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    text += "\t".join(str(cell) if cell is not None else "" for cell in row) + "\n"
        else:  # .xls
            df = pd.read_excel(file_path, sheet_name=None)
            for sheet_name, data in df.items():
                text += f"\n\nSheet: {sheet_name}\n{data.to_string(index=False)}\n"

        return {
            "content": [{
                "page": 1,
                "text": text,
                "type": "table_data"
            }]
        }
    except Exception as e:
        logger.error(f"Excel processing failed: {str(e)}")
        raise

def process_image(file_path):
    """Extract text from image file (JPG, JPEG, PNG)"""
    try:
        image = Image.open(file_path)
        text = extract_text_from_image(image)
        return {
            "content": [{
                "page": 1,
                "text": text,
                "type": "ocr_text"
            }]
        }
    except Exception as e:
        logger.error(f"Image processing failed: {str(e)}")
        raise

@app.route('/process', methods=['POST'])
@auth.login_required
def handle_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    if not allowed_file(file.filename):
        return jsonify({"error": "Invalid file type"}), 400

    temp_path = None
    try:
        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, filename)
        file.save(temp_path)

        start_time = time.time()
        file_extension = filename.rsplit('.', 1)[1].lower()

        # Process file based on extension
        if file_extension == 'pdf':
            results = []
            with pdfplumber.open(temp_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    page_result = process_pdf_page(page, page_num, temp_path)
                    results.append(page_result)

            # Combine results
            combined_text = ""
            for page in results:
                combined_text += page.get("native_text", "") + "\n" + page.get("image_text", "") + "\n"

            response = {
                "metadata": {
                    "filename": filename,
                    "pages": len(results),
                    "processing_time": round(time.time() - start_time, 2),
                    "text_length": len(combined_text)
                },
                "content": results
            }
        elif file_extension == 'docx':
            response = process_docx(temp_path)
            response['metadata'] = {
                "filename": filename,
                "pages": 1,
                "processing_time": round(time.time() - start_time, 2),
                "text_length": len(response['content'][0]['text'])
            }
        elif file_extension == 'txt':
            response = process_txt(temp_path)
            response['metadata'] = {
                "filename": filename,
                "pages": 1,
                "processing_time": round(time.time() - start_time, 2),
                "text_length": len(response['content'][0]['text'])
            }
        elif file_extension == 'csv':
            response = process_csv(temp_path)
            response['metadata'] = {
                "filename": filename,
                "pages": 1,
                "processing_time": round(time.time() - start_time, 2),
                "text_length": len(response['content'][0]['text'])
            }
        elif file_extension in ('xlsx', 'xls'):
            response = process_excel(temp_path)
            response['metadata'] = {
                "filename": filename,
                "pages": 1,
                "processing_time": round(time.time() - start_time, 2),
                "text_length": len(response['content'][0]['text'])
            }
        elif file_extension in ('jpg', 'jpeg', 'png'):
            response = process_image(temp_path)
            response['metadata'] = {
                "filename": filename,
                "pages": 1,
                "processing_time": round(time.time() - start_time, 2),
                "text_length": len(response['content'][0]['text'])
            }
        else:
            return jsonify({"error": "Unsupported file type"}), 400

        return jsonify(response)

    except Exception as e:
        logger.error(f"Processing failed: {str(e)}")
        return jsonify({"error": str(e)}), 500

    finally:
        # Clean up temporary files
        try:
            if temp_path and os.path.exists(temp_path):
                os.remove(temp_path)
            if 'temp_dir' in locals() and os.path.exists(temp_dir):
                os.rmdir(temp_dir)
        except Exception as e:
            logger.error(f"Cleanup failed: {str(e)}")

handler = app


